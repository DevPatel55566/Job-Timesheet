from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
EXCEL_FILE = 'timesheet.xlsx'

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Week", "Date", "Sign In", "Sign Out", "Total Hours", "Break", "Payment Received"])
    wb.save(EXCEL_FILE)

def read_all_rows():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

@app.route('/')
def index():
    rows = read_all_rows()
    return render_template('index.html', entries=rows, message=request.args.get('msg'))

@app.route('/submit', methods=['POST'])
def submit():
    data = [
        request.form['week'],
        request.form['date'],
        request.form['signin'],
        request.form['signout'],
        request.form['hours'],
        request.form['break'],
        request.form['payment']
    ]

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(data)
    wb.save(EXCEL_FILE)

    return redirect(url_for('index', msg='Form submitted successfully!'))


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host='0.0.0.0', port=port)